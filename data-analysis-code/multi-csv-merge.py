#   -*- coding:utf-8 -*-
"""
# @Author: Magician
# @Date: 2021-01-14 18:24:10
# @Description: 合并多个csv文件到一个excel表格中

# Copyright 2020 by Magician
"""
import openpyxl  # 读写.xlsx文件
import pandas as pd
import csv
import glob
import os
pd.set_option('display.max_columns',1000)
# pd.set_option('display.max_rows',n)#最多显示的行数
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth',1000)
root_path = 'C:/Users/21332/test/ce_feature'  #这里填已经存储了各个csv文件的表格

min_line = 21
max_line = 885


dst_filepath = 'C:/Users/21332/test/ce_feature.xlsx'   #这里是导出来的表格的路径
dst = openpyxl.Workbook()
# 这一步是用来简化后面的路径的，类似pandas as pd

dst_sheet = dst.active
# 新建workbook：通过openpyxl.workbook.Workbook.active()得到worksheet
dst_sheet.title="Sheet1"

csv_list = glob.glob(root_path+'/*.csv')
# 这一步是用来读取文件夹里面的CSV列表的
# glob.glob(pathname)：返回匹配 pathname 的路径名列表，其中的元素必须为包含一个路径信息的字符串。
# 返回的是列表 list类型。是所有路径下的符合条件的文件名的列表。
print(u'共发现%s个CSV文件'%len(csv_list))
print(u'正在处理…')

write_value = [] # 创建write_value这么一个列表

for csv_file in csv_list: #循环读取同文件夹下的csv文件
    with open(csv_file,'r',encoding='utf-8') as csvfile:
        # Python文件操作中的读写模式:open(path, '-模式-',encoding='UTF-8')，'r'是“只读”的意思
        reader = csv.reader(csvfile) #csv.reader（）读取结果是列表
        column=[row[3] for row in reader] # 遍历每一行的第4个数据，读取的是csv里对应特征的数值
        write_value.append(column[23:884]) # python中的append()方法用于在列表末尾添加新的对象。
        '''for csv_index,rows in enumerate(reader):
           if csv_index in range(min_line-1,max_line+1):
              write_value.append(rows)'''

for csv_file in csv_list: #循环读取同文件夹下的csv文件
    id = csv_file.replace(root_path,'').replace('.csv','')
    csv_num = csv_list.index(csv_file)
    print(csv_num)
    dst_sheet.cell(csv_num+1,1,id) # 单元格的赋值：sheet.cell(row=3,column=5,value=100)，row/column表示被赋值的单元格的坐标。
    # 这一步是把患者姓名写入表格的第一列
    #print(write_value[1][20])
    for j in range(len(write_value[0])):
        # print(write_value[csv_num][j])
        row=csv_num+1
        column1=j+2
        dst_sheet.cell(row, column1).value=write_value[csv_num][j]
    
    #for i in range(line_num):
     #   for j in range(4):
      #      dst_sheet.cell(line_num + 1, i + 2, write_value[csv_num][j])

dst.save(filename=dst_filepath)   
print(u'合并完毕！')
